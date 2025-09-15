import json
import time
from openpyxl import load_workbook

def excel_to_json(file_path, sheet_name=0):
    wb = load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))

    # 第一行：注释（headers）
    #headers = list(rows[0])
    # 第三行：数据类型（types）
    types = list(rows[2])
    # 第四行：字段名（keys）- 这里我们只需要顺序，不用于对象键
    keys = list(rows[3])

    # 第五行及之后：数据
    data = []
    for row in rows[4:]:
        data.append(list(row))

    result = {
        "headers": keys,
        "types": types,
        "data": data
    }

    return result

# 使用示例
if __name__ == "__main__":
    file_path = "../example/Z.资源.xlsx"  # 替换为你的 Excel 文件路径
    
    # 开始计时
    start_time = time.time()
    
    result = excel_to_json(file_path)

    # 输出到控制台
    # print(json.dumps(result, ensure_ascii=False, indent=2))

    # 或写入文件
    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    # 结束计时并打印执行时间
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"执行时间: {execution_time:.4f} 秒")